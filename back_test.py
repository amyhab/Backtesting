# -*- coding: utf-8 -*-
"""
Created on Thu Jan 04 09:44:01 2018
Function：backtest by stock seleced by factor from target stock pool
@author: amyhab
"""

import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages  
import numpy as np
import pandas as pd
from scipy import stats
from openpyxl import Workbook  
from openpyxl.writer.excel import ExcelWriter
import MySQLdb
from mosek.fusion import Model,Domain,Expr,ObjectiveSense

class MSSQL(object):
    """
       use SQL to get market data from database
    """
    def __init__(self,HOST = '192.168.1.1',PORT = 3311,DATABASE = 'database',USER = 'amyhab',PASSWORD = 'amyhab',CHARSET = "utf8"):
        self.host = HOST
        self.port = PORT
        self.db = DATABASE
        self.user = USER
        self.pwd = PASSWORD
        self.charset = CHARSET
    def __GetConnect(self):
        if not self.db:
            raise(NameError,"No Database information")
        self.conn = MySQLdb.connect(host=self.host,port=self.port,user=self.user,passwd=self.pwd,db=self.db,charset = self.charset)
        cur = self.conn.cursor()
        if not cur:
            raise(NameError,"Connecting Database failed")
        else:
            return cur
    def exesql(self,sql,param = None):
        cur = self.__GetConnect()
        if param == None:
            cur.execute(sql)
        else:
            cur.execute(sql,param)
        result = cur.fetchall()
        self.conn.commit()
        self.conn.close()
        return result

def trade_day(ms,start,end):
    # read Trade Date
    sql='''
SELECT CALENDAR_DATE FROM vsec_trade_cal WHERE IS_OPEN= 1 And CALENDAR_DATE BETWEEN '%s' AND '%s' AND EXCHANGE_CD='XSHG' '''%(start,end)       
    try:
        sql_result = ms.exesql(sql)
    except:
        raise(NameError,'exesql error: The input parameter of trade_day may not be correct.')
    result = []
    for i in xrange(len(sql_result)):
        sqli = sql_result[i][0].strftime("%Y-%m-%d")
        result.append(sqli)
    return result

def getdata_tradeday(ms,start,end):
    # generate dict of trade date for query
    Date = trade_day(ms,start,end)
    dict_date = {}
    for i in xrange(len(Date)):
        dict_date.update({Date[i]:i})
    return Date,dict_date

def CSI_info(ms,end,code):
    # read consitution stock of stock pool,return [code,into date, out date]
    sql_result = []
    sql = '''
SELECT CONS_TICKER_SYMBOL,INTO_DATE,OUT_DATE FROM vidx_cons WHERE TICKER_SYMBOL='%s' order by CONS_TICKER_SYMBOL''' %(code)  
    try:
        sql_result = ms.exesql(sql)
    except:
        raise(NameError,'exesql error: The input parameter of stock_info may not be correct.')
    result = []
    for i in xrange(len(sql_result)):
        if sql_result[i][2]!=None:
            sqli = [str(sql_result[i][0]),sql_result[i][1].strftime("%Y-%m-%d"),sql_result[i][2].strftime("%Y-%m-%d")]
        else:
            sqli = [str(sql_result[i][0]),sql_result[i][1].strftime("%Y-%m-%d"),end]            
        result.append(sqli)
    return result
    
def getdata_CSI(ms,endday,code):
    # generate dict of codelist for query
    CSI = CSI_info(ms,endday,code)
    # note: one stock may be listed in and then delisted from index for many times
    codelist = []
    dict_code = {}
    for i in xrange(len(CSI)):
        if not dict_code.has_key(str(CSI[i][0])):
            codelist.append(str(CSI[i][0]))
            dict_code.update({str(CSI[i][0]):len(codelist)-1})
    return CSI,codelist,dict_code

def get_merger(ms,code,end):
    # get delist information 
    # note: some stocks may be merged and delisted before they are delisted from index
    sql_result = []
    sql = '''
SELECT TICKER_SYMBOL,DELIST_DATE FROM vequ WHERE TICKER_SYMBOL in %s order by TICKER_SYMBOL'''%(code,)   
    try:
        sql_result = ms.exesql(sql)
    except:
        raise(NameError,'exesql error: The input parameter of get_merger may not be correct.')
    result = []
    for i in xrange(len(sql_result)):
        if sql_result[i][1]!=None:
            sqli = [str(sql_result[i][0]),sql_result[i][1].strftime("%Y-%m-%d")]
        else:
            sqli = [str(sql_result[i][0]),end]            
        result.append(sqli)
    return result

def getdata_merger(ms,codelist,startday,endday,stock_info,dict_date,dict_code):
    # some stocks may be merged and delisted before they are delisted from index
    # note: we suppose we know the information 120 trade day before the delist day
    delist_date = get_merger(ms,tuple(codelist),endday)
    out_date = [startday]*len(delist_date)
    for j in xrange(len(stock_info)):
        pos_code = dict_code[stock_info[j][0]]
        if stock_info[j][2]>out_date[pos_code]:
            out_date[pos_code] = stock_info[j][2]    
    is_merger = [[0,0] for i in xrange(len(codelist))]
    for i in xrange(len(codelist)):
        if delist_date[i][1]<=out_date[i] and delist_date[i][1]<endday:
            is_merger[i][0]=1
            is_merger[i][1]=dict_date[delist_date[i][1]]-120
    return is_merger
    
def mkt_stock(ms,code,start,end):
    # read highest price,lowest price,pre close price,close price,turnover to calculate VWAP and buy/sell condition
    sql_result = []
    sql = '''
SELECT TRADE_DATE,TICKER_SYMBOL,SEC_SHORT_NAME,HIGHEST_PRICE,LOWEST_PRICE,PRE_CLOSE_PRICE,CLOSE_PRICE,TURNOVER_VOL,TURNOVER_VALUE FROM vmkt_equd WHERE TICKER_SYMBOL in %s AND TRADE_DATE BETWEEN '%s' AND '%s' order by TICKER_SYMBOL, TRADE_DATE'''%(code,start,end)                         
    try:
        sql_result = ms.exesql(sql)
    except:
        raise(NameError,'exesql error: The input parameter of mkt_stock may not be correct.')
    result = []
    for i in xrange(len(sql_result)):
        sqli = [str(sql_result[i][1]),sql_result[i][0].strftime("%Y-%m-%d"),sql_result[i][2]]
        for j in xrange(6):
            sqli.append(float(sql_result[i][j+3]))
        result.append(sqli)
    return result

def data_process(mktdata,Date,codelist,dict_date,dict_code):
    # generate 4 Dataframe: pct_chg from preclose to vwap, pct_chg from vwap to close, if we can buy today, if we can sell today
    len_stock = len(codelist)
    len_Date = len(Date)
    pct_chg1 = [[-1,]*len_Date for x in xrange(len_stock)]
    pct_chg2 = [[-1,]*len_Date for x in xrange(len_stock)]
    buy_if = [[-1,]*len_Date for x in xrange(len_stock)]
    sell_if = [[-1,]*len_Date for x in xrange(len_stock)]
    for i in xrange(len(mktdata)):
        tdate = mktdata[i][1]
        codei = mktdata[i][0]
        pos_date = dict_date[tdate]
        pos_code = dict_code[codei]
        high = mktdata[i][3]
        low = mktdata[i][4]
        pre_close = mktdata[i][5]
        close = mktdata[i][6]
        volume = mktdata[i][7]
        volvalue = mktdata[i][8]
        # not suspension
        # note: since we select stocks from index, there is no new or ST stock
        if volume<>0 :
            # not price limits when open
            if high<>low:
                if high>=pre_close*1.097:
                # note: suppose we can trade half when there is price limits intraday
                    buy_if[pos_code][pos_date]=0
                else:
                    buy_if[pos_code][pos_date]=1
                if low<=pre_close*0.903:
                    sell_if[pos_code][pos_date]=0
                else:
                    sell_if[pos_code][pos_date]=1
            elif high>=pre_close:
                # buy price limits:can sell
                sell_if[pos_code][pos_date]=1
            elif low<=pre_close:
                # sell price limits:can buy
                buy_if[pos_code][pos_date]=1
            vwap = volvalue/volume
            pctchg1 = (vwap-pre_close)/pre_close
            pctchg2 = (close-vwap)/vwap
            pct_chg1[pos_code][pos_date]=pctchg1
            pct_chg2[pos_code][pos_date]=pctchg2
    pct_chg1 = pd.DataFrame(pct_chg1,index=codelist,columns=Date)
    pct_chg2 = pd.DataFrame(pct_chg2,index=codelist,columns=Date)
    buy_if = pd.DataFrame(buy_if,index=codelist,columns=Date)
    sell_if = pd.DataFrame(sell_if,index=codelist,columns=Date)
    return pct_chg1,pct_chg2,buy_if,sell_if   
    
def mkt_index(ms,code,start,end):
    # read pre close price,close price of index
    sql_result = []
    sql = '''
SELECT TRADE_DATE,TICKER_SYMBOL,PRE_CLOSE_INDEX,CLOSE_INDEX FROM vmkt_idxd WHERE TICKER_SYMBOL= %s AND TRADE_DATE BETWEEN '%s' AND '%s' order by TICKER_SYMBOL, TRADE_DATE'''%(code,start,end)                         
    try:
        sql_result = ms.exesql(sql)
    except:
        raise(NameError,'exesql error: The input parameter of mkt_index may not be correct.')
    result = []
    for i in xrange(len(sql_result)):
        sqli = [str(sql_result[i][1]),sql_result[i][0].strftime("%Y-%m-%d"),float(sql_result[i][2]),float(sql_result[i][3])]
        result.append(sqli)
    return result


def indus_info(ms,code,date,endday):
    # read SW 1st-industry category information where short name used for modifying bank and nonbank
    sql_result = []
    sql = '''
SELECT TICKER_SYMBOL,SEC_SHORT_NAME,INDUSTRY_ID_1ST,INTO_DATE,OUT_DATE FROM vequ_industry WHERE TICKER_SYMBOL in %s AND INDUSTRY_VERSION_CD='010303' AND (OUT_DATE>='%s' OR  IS_NEW=1) AND (INTO_DATE<='%s') order by TICKER_SYMBOL'''%(code,date,endday)  
    try:
        sql_result = ms.exesql(sql)
    except:
        raise(NameError,'exesql error: The input parameter of indus_info may not be correct.')
    result = []
    for i in xrange(len(sql_result)):
        sqli = [str(sql_result[i][0]),sql_result[i][1],str(sql_result[i][2]),sql_result[i][3].strftime("%Y-%m-%d")]
        if sql_result[i][4]!=None:
            sqli.append(sql_result[i][4].strftime("%Y-%m-%d"))
        else:
            sqli.append(endday)
        result.append(sqli)
    return result

def base_indus_info(ms,code,date):
    # read SW 1st-industry category information at specific time point
    sql_result = []
    sql = '''
SELECT TICKER_SYMBOL,INDUSTRY_ID_1ST FROM vequ_industry WHERE TICKER_SYMBOL in %s AND INDUSTRY_VERSION_CD='010303' AND INTO_DATE<='%s' AND (OUT_DATE>='%s' OR  IS_NEW=1) order by TICKER_SYMBOL'''%(code,date,date)  
    try:
        sql_result = ms.exesql(sql)
    except:
        raise(NameError,'exesql error: The input parameter of base_indus_info may not be correct.')
    result = []
    for i in xrange(len(sql_result)):
        sqli = [str(sql_result[i][0]),str(sql_result[i][1])]
        result.append(sqli)
    return result

def get_indus(ms,codelist,Date,startday,endday,base_date,dict_date,dict_code):
    # get stock industry category data 
    # note: there is no bank and nonbank category before 2014.1.1, use baseday information to modify
    category_indus = []
    rawdata = indus_info(ms,tuple(codelist),startday,endday)
    print('indus data loaded')
    base_indus = base_indus_info(ms,tuple(codelist),base_date)
    len_stock = len(codelist)
    len_Date = len(Date)
    indus= [['',]*len_Date for x in xrange(len_stock)]
    for i in xrange(len(rawdata)):
        tdate = rawdata[i][3]
        # modify change date into tradedays
        for j in xrange(len(Date)-1):
            if tdate<Date[0]:
                tdate = Date[0]
                break
            elif tdate>Date[j] and tdate<Date[j+1]:
                tdate = Date[j+1]
                break
            elif tdate==Date[j]:
                break
        codei = rawdata[i][0]
        value = rawdata[i][2]
        pos_date = dict_date[tdate]
        pos_code = dict_code[codei]
        # 01030333 means Finance in SW category before 2014, need to specify
        if tdate<'2014-01-01' and value=='01030333':
            flag = True
            match_data = ''
            for k in xrange(len(base_indus)):
                if codei==base_indus[k][0]:
                    flag = False
                    match_data = base_indus[k][1]
                    break
            # match data with baseday info
            if (not flag) and (match_data=='01030321' or match_data=='01030322'):
                value=match_data
            # cannot match: use sec_short_name
            elif rawdata[i][1][2:4]==u'银行':
                value='01030321'
            else:
                value='01030322'
        indus[pos_code][pos_date] = value  
    # backward assignment
    for i in xrange(len(codelist)):
        for j in xrange(len(Date)-1):
            if indus[i][j]<>'' and indus[i][j+1]=='':
                indus[i][j+1]=indus[i][j]
    for j in xrange(len(Date)):#确定当日行业种类
        category_j = []
        for i in xrange(len(codelist)):
            if (indus[i][j] not in category_j) and (indus[i][j]<>''):
                category_j.append(indus[i][j])
        category_indus.append(category_j)
    return indus,category_indus

def index_weight_info(ms,index_code,startday,endday,Date):
    # read weight of consitution stock of index
    sql_result = []
    sql = '''
SELECT CONS_TICKER_SYMBOL,EFF_DATE,WEIGHT FROM vidx_trad_sh WHERE TICKER_SYMBOL='%s' AND DATA_TYPE=1 AND EFF_DATE in %s order by EFF_DATE,CONS_TICKER_SYMBOL'''%(index_code,Date)  
    try:
        sql_result = ms.exesql(sql)
    except:
        raise(NameError,'exesql error: The input parameter of index_weight may not be correct.')
    result = []
    for i in xrange(len(sql_result)):
        if sql_result[i][2]!=None:
            sqli = [str(sql_result[i][0]),sql_result[i][1].strftime("%Y-%m-%d"),float(sql_result[i][2])]
        else:
            sqli = [str(sql_result[i][0]),sql_result[i][1].strftime("%Y-%m-%d"),0]
        result.append(sqli)
    return result

def index_weight(ms,codelist,Date,index_code,startday,endday,indus,dict_code):
    # calculate bank, nonbank and realestate weight of index
    rawdata = index_weight_info(ms,index_code,startday,endday,tuple(Date))
    weight_bank = []
    weight_nonbank = []
    weight_realestate = []
    i = 0
    for j in xrange(len(Date)):
        tdate = Date[j]
        dweight_bank = 0
        dweight_nonbank = 0
        dweight_realestate = 0
        while i<len(rawdata) and rawdata[i][1]==tdate:
            codei = rawdata[i][0]
            weight = rawdata[i][2]
            pos_code = dict_code[codei]
            if indus[pos_code][j]=='01030321':
                dweight_bank += weight
            elif indus[pos_code][j]=='01030322':
                dweight_nonbank += weight
            elif indus[pos_code][j]=='01030320':
                dweight_realestate += weight
            i = i + 1
        weight_bank.append(dweight_bank)
        weight_nonbank.append(dweight_nonbank)
        weight_realestate.append(dweight_realestate)
    return weight_bank,weight_nonbank,weight_realestate

        
class Backtest(object):
    """Backtest plantform
    Parameters：
    ----------
    index:str,default = '000906'
        stock pool
    read_db_if:int,default = 1
        read market data from database or local, where 1 from local
    portfolio_num:int,default = 60
        stock number of target portfolio
    hedge_target:str,default = 'half'
        stock future used for hedging, must be consistent with index
        '000300':use CSI300
        '000905':use CSI905
        'half':use mixture of CSI300 and CSI905 by weight=0.5
    factor_order:int,default = 1
        whether to reverse the factor order
        0:ascending order
        1:descending order
    trade_cost:float,default = 0.0012
        trade cost or impact cost
    industry_control:int,default = 1
        0:not control industry weight
        1:control bank and nonbank
        2:control bank,nonbanka and real_estate
    weight_low_bank:float,default = 0.0
        max low weight from index weight for industry bank,meaningless if not control bank
    weight_low_nonbank:float,default = 0.0
        max low weight from index weight for industry nonbank,meaningless if not control nonbank
    weight_low_realestate:float,default = 0.0
        max low weight from index weight for industry realestate,meaningless if not control realestate
    factor_standardize:int,default = 0
        0:not do standardize for factor
        1:standardize for factor
    industry_neutral:int,default = 0
        0:not do industry neutralize for factor
        1:industry neutralize for factor
    portfolio_weight:str,default = 'Equal'
        How to calculate stock weight in target portfolio
        'Equal':equal weight
        'Factor':weight due to factor
        'Mosek':weight due to Mosek, in this method portfolio_num just used to open position
    mosek_high_bound:float,default = 0.02
        only used when portfolio_weight = 'Mosek', set the max stock weight to form target portfolio
    ic_day:int,default = 20
        ic_day's tradeday percent change used to calculate ic
    startday:str
        start date for backtesting,must be consistent with market data
    endday:str
        end date for backtesting,must be consistent with market data
    """
    def __init__(self,index='000906',read_db_if=1,portfolio_num=60,hedge_target='half',factor_order=1,trade_cost=0.0012,
                 industry_control=1,weight_low_bank=0.0,weight_low_nonbank=0.0,weight_low_realestate=0.0,
                 factor_standardize=0,industry_neutral=0,portfolio_weight='Equal',mosek_high_bound=0.02,ic_day=20):
        self.index = index
        self.read_db_if = read_db_if
        self.portfolio_num = portfolio_num
        self.hedge_target = hedge_target
        self.factor_order = factor_order
        self.trade_cost = trade_cost
        self.industry_control = industry_control
        self.weight_low_bank = weight_low_bank
        self.weight_low_nonbank = weight_low_nonbank
        self.weight_low_realestate = weight_low_realestate
        self.factor_standardize = factor_standardize
        self.industry_neutral = industry_neutral
        self.portfolio_weight = portfolio_weight
        self.mosek_high_bound = mosek_high_bound
        self.ic_day = ic_day
        self.net_value = []
        self.trade_date = []
        self.turnover_rate = []
        self.actual_portfolio_num = []
        self.weight_300 = []
        self.rank_ic = []
        self.daily_excess_return = []
    
    
    def valuate(self,excess_return,tdate):
        """
        calculate some indicator to measure performance of strategy
        note: use 250 trade day as a year
              since we use addition for net value, max drawdown must be subtraction
        """
        netvalue = 1
        car = []
        for i in xrange(len(excess_return)):
            netvalue = netvalue + excess_return[i]
            car.append(netvalue)
        netvalue = car[len(car)-1]
        year = float(len(car))/250
        # annualized reutrn
        annual_r = (netvalue-1)/year#年化收益
        x = np.array(excess_return)
        std = np.sqrt(250)*np.std(x)
        # sharpe ratio
        sharpe_ratio = annual_r/std
        # max drawdown and its continuing days
        drawdown = 0
        drawdownday = 0
        for i in xrange(1,len(car)-1):
            temp =car[0:i]
            y = np.array(temp)
            maxn = np.max(y)
            maxpos = np.argmax(y)
            down = (maxn-car[i])
            if (down>drawdown):
                drawdown = down
                drawdownday = maxpos-i
        # min excess return and its date
        daymin = np.min(excess_return)
        minday = tdate[np.argmin(excess_return)]
        winnum = 0
        profit = 0
        loss = 0
        for i in xrange(len(excess_return)):
            if excess_return[i]>=0:
                winnum += 1
                profit += excess_return[i]
            else:
                loss += np.abs(excess_return[i])
        # winning rate
        winratio = float(winnum)/len(excess_return)
        # profit loss ratio        
        p2l = profit/loss    
        re = '%.2f'%(annual_r*100)
        mdd = '%.2f'%(drawdown*100)
        mdds = str(drawdownday)
        sh = '%.2f'%(sharpe_ratio)
        dm = '%.2f'%(daymin*100)
        wr = '%.2f'%(winratio)
        pl = '%.2f'%(p2l)
        result = [re,mdd,mdds,sh,dm,minday,wr,pl]
        return result  
    
    def data_update(self,ms,codelist,startday,endday,Date):
        mktdata_daily = mkt_stock(ms,tuple(codelist),startday,endday)
        print(u'market data loaded')    
        pct_chg1,pct_chg2,buy_if,sell_if = data_process(mktdata_daily,Date,codelist)
        print(u'pct_chg calculated')
        pct_chg1.to_csv('.\\bdata\\pct1.csv')
        pct_chg2.to_csv('.\\bdata\\pct2.csv')
        buy_if.to_csv('.\\bdata\\buy.csv')
        sell_if.to_csv('.\\bdata\\sell.csv')
        return pct_chg1,pct_chg2,buy_if,sell_if
        
    def get_backtest_data(self,startday,endday,base_indus_day):
        """
        get data for backtesting:
        Date: trade dates and dict of it
        stock_info: stock name,listed date and delisted date
        codelist: all stock once be listed in stock pool and dict of it
        is_merger: whether the stock be merged and supposed not buy day
        pct_chg1: percent change from preclose to vwap
        pct_chg2: percent change from vwap to close
        buy_if: if we can buy today
        sell_if: if we can sell today
        Date_Match: if local data matches Date
        indus: industry category for stock each trade day
        category_indus: each trade day's all industry category
        index_info: percent change, weight of bank, nonbank and realestate of hedge target
        """
        # Date_Match used to delect Data accuracy
        Date_Match = True
        ms = MSSQL()
        Date,dict_date = getdata_tradeday(ms,startday,endday)
        stock_info,codelist,dict_code = getdata_CSI(ms,endday,self.index) 
        is_merger = getdata_merger(ms,codelist,startday,endday,stock_info,dict_date,dict_code)
        if self.read_db_if==0:
            print 'update data from database'
            pct_chg1,pct_chg2,buy_if,sell_if = self.data_update(ms,codelist,startday,endday,Date,dict_date,dict_code)
        else:
            pct_chg1 = pd.read_csv('.\\bdata\\pct1.csv',index_col=0)
            pct_chg1 = pct_chg1.values.tolist()
            print('pct1 loaded')
            pct_chg2 = pd.read_csv('.\\bdata\\pct2.csv',index_col=0)
            pct_chg2 = pct_chg2.values.tolist()
            print('pct2 loaded')
            buy_if = pd.read_csv('.\\bdata\\buy.csv',index_col=0)
            dbDate = [x for x in buy_if.columns]
            buy_if = buy_if.values.tolist()
            print('buy_if loaded')
            sell_if = pd.read_csv('.\\bdata\\sell.csv',index_col=0)
            sell_if = sell_if.values.tolist()
            print('sell_if loaded')
            if dbDate[0]<>Date[0] or dbDate[len(dbDate)-1]<>Date[len(Date)-1]:
                Date_Match = False
        factor = pd.read_csv('.\\bdata\\factor.csv',index_col=0)
        factor.index = codelist
        factorDate = [x for x in factor.columns]
        if factorDate[0]<>Date[0] or factorDate[len(factorDate)-1]<>Date[len(Date)-1]:
            Date_Match = False
        factor = factor.values.tolist()
        print('factor loaded')
        indus = []
        category_indus = []
        index_info = []
        if self.industry_control<>0:
            # control industry weight
            indus,category_indus = get_indus(ms,codelist,Date,startday,endday,base_indus_day,dict_date,dict_code)  
            print (u'category information loaded')  
            if self.hedge_target=='000300':
                weight_bank300,weight_nonbank300,weight_realestate300 = index_weight(ms,codelist,Date,self.hedge_target,startday,endday,indus,dict_code)
                index_300 = mkt_index(ms,self.hedge_target,startday,endday) 
                for i in xrange(len(index_300)):
                    row = [index_300[i][1],index_300[i][3]/index_300[i][2]-1,weight_bank300[i],weight_nonbank300[i],weight_realestate300[i]]
                    index_info.append(row)
            elif self.hedge_target=='000905':
                weight_bank905,weight_nonbank905,weight_realestate905 = index_weight(ms,codelist,Date,self.hedge_target,startday,endday,indus,dict_code)
                index_905 = mkt_index(ms,self.hedge_target,startday,endday) 
                for i in xrange(len(index_905)):
                    row = [index_905[i][1],index_905[i][3]/index_905[i][2]-1,weight_bank905[i],weight_nonbank905[i],weight_realestate905[i]]
                    index_info.append(row)
            elif self.hedge_target=='half':
                weight_bank300,weight_nonbank300,weight_realestate300 = index_weight(ms,codelist,Date,self.hedge_target,startday,endday,indus,dict_code)
                index_300 = mkt_index(ms,self.hedge_target,startday,endday) 
                weight_bank905,weight_nonbank905,weight_realestate905 = index_weight(ms,codelist,Date,self.hedge_target,startday,endday,indus,dict_code)
                index_905 = mkt_index(ms,self.hedge_target,startday,endday) 
                for i in xrange(len(index_300)):
                    row = [index_300[i][1],0.5*(index_300[i][3]/index_300[i][2]+index_905[i][3]/index_905[i][2])-1,\
                    0.5*(weight_bank300[i]+weight_bank905[i]),0.5*(weight_nonbank300[i]+weight_nonbank905[i]),\
                    0.5*(weight_realestate300[i]+weight_realestate905[i])]
                    index_info.append(row)
        else:
            # not control industry weight
            if self.hedge_target=='000300':
                index_300 = mkt_index(ms,self.hedge_target,startday,endday) 
                for i in xrange(len(index_300)):
                    row = [index_300[i][1],index_300[i][3]/index_300[i][2]-1,0,0,0]
                    index_info.append(row)
            elif self.hedge_target=='000905':
                index_905 = mkt_index(ms,self.hedge_target,startday,endday) 
                for i in xrange(len(index_905)):
                    row = [index_905[i][1],index_905[i][3]/index_905[i][2]-1,0,0,0]
                    index_info.append(row)
            elif self.hedge_target=='half':
                index_300 = mkt_index(ms,self.hedge_target,startday,endday) 
                index_905 = mkt_index(ms,self.hedge_target,startday,endday) 
                for i in xrange(len(index_300)):
                    row = [index_300[i][1],0.5*(index_300[i][3]/index_300[i][2]+index_905[i][3]/index_905[i][2])-1,0,0,0]
                    index_info.append(row)
        return Date_Match,Date,dict_date,dict_code,stock_info,codelist,dict_code,is_merger,pct_chg1,pct_chg2,buy_if,sell_if,factor,indus,category_indus,index_info
           
           
           
    def cal_tweight(self,actual_data,w_bank,w_nonbank,w_realestate,indus,pos_date,pct_chg1,pct_chg2,Date,dict_code):
        """
        calculate target portfolio(stock and weight) after one tradeday close
        Constraint1: different weight calculating method
        Constraint2: diiferent industry control method
        Constraint3: Second Board weight less than 0.2
        """
        t_factor = [actual_data[i][1] for i in xrange(len(actual_data))]
        # factor standardize, especially 3sigma outlier
        if self.factor_standardize==1:
            mu = np.average(t_factor)
            sigma = np.std(t_factor)
            unit_factor = []
            for i in xrange(len(t_factor)):
                if t_factor[i]>(mu-3*sigma) and t_factor[i]<(mu+3*sigma):
                    unit_factor.append(t_factor[i])
            mu = np.average(unit_factor)
            sigma = np.std(unit_factor)
            for i in xrange(len(t_factor)):
                if t_factor[i]>(3*sigma+mu):
                    t_factor[i]=3
                elif t_factor[i]<(-3*sigma+mu):
                    t_factor[i]=-3
                else:
                    t_factor[i] = (t_factor[i]-mu)/sigma
        if pos_date<(len(Date)-self.ic_day):  
            pct_chg = []
            for i in xrange(len(actual_data)):
                codei = actual_data[i][0]
                pos_code = dict_code[codei]
                pct = 1
                for j in xrange(pos_date,pos_date+self.ic_day):
                    pct1 = pct_chg1[pos_code][j]
                    pct2 = pct_chg2[pos_code][j]
                    if pct1==-1:
                        pct1 = 0
                    if pct2==-1:
                        pct2 = 0
                    pct = pct*(1+pct1)*(1+pct2)
                pct_chg.append(pct)
            ic,pval = stats.spearmanr(t_factor,pct_chg)
        else:
            ic = 0
        # different portfolio weight method 
        # note: Mosek used to solve different constraints
        if self.portfolio_weight=='Mosek':
            result = []
            set_portfolio = []
            A_bank = [0 for i in xrange(len(actual_data))]
            A_nonbank = [0 for i in xrange(len(actual_data))]
            A_realestate = [0 for i in xrange(len(actual_data))]
            A_300 = [0 for i in xrange(len(actual_data))]
            num_bank = 0
            num_nonbank = 0
            num_realestate = 0
            for j in xrange(len(actual_data)):
                codej = actual_data[j][0]
                pos_code = dict_code[codej]
                if codej[0:3]=='300':# second board
                    A_300[j] = 1
                if indus[pos_code][pos_date-1]=='01030321':# bank
                    A_bank[j] = 1
                    num_bank += 1 
                elif indus[pos_code][pos_date-1]=='01030322':# nonbank
                    A_nonbank[j] = 1
                    num_nonbank += 1 
                elif indus[pos_code][pos_date-1]=='01030320':# realestate
                    A_realestate[j] = 1
                    num_realestate += 1
            n = len(actual_data)
            with  Model("Weight Optimize") as M:
                x = M.variable("x", n, Domain.greaterThan(0.0))
                if self.factor_order==1:
                    M.objective("obj", ObjectiveSense.Maximize, Expr.dot(t_factor, x))
                else:
                    M.objective("obj", ObjectiveSense.Minimize, Expr.dot(t_factor, x))
                # note: In case no solution, set industry constraint when need and possible(enough industry stocks in stock pool)
                if (num_bank*self.mosek_high_bound>=w_bank) and (num_nonbank*self.mosek_high_bound>=w_nonbank) and (num_realestate*self.mosek_high_bound>=w_realestate):
                    M.constraint('maxx', x, Domain.lessThan(self.mosek_high_bound))
                else:
                    for i in xrange(len(actual_data)):
                        # whether control bank stock weight
                        if A_bank[i]==1 and num_bank<>0 and self.industry_control<>0: 
                            M.constraint(str(i),x.index(i),Domain.lessThan(max(self.mosek_high_bound,w_bank/num_bank)))
                        # whether control nonbank stock weight
                        elif A_nonbank[i]==1 and num_nonbank<>0 and self.industry_control<>0:
                            M.constraint(str(i),x.index(i),Domain.lessThan(max(self.mosek_high_bound,w_nonbank/num_nonbank)))
                        # whether control realestate stock weight
                        elif A_realestate[i]==1 and num_realestate<>0 and self.industry_control==2:
                            M.constraint(str(i),x.index(i),Domain.lessThan(max(self.mosek_high_bound,w_realestate/num_realestate)))
                        else:
                            M.constraint(str(i),x.index(i),Domain.lessThan(self.mosek_high_bound))
                M.constraint('budget', Expr.sum(x), Domain.equalsTo(1.0))
                # whether control bank industry weight
                if (not (w_bank>0 and num_bank==0)) and self.industry_control<>0:
                    M.constraint('bank', Expr.dot(A_bank, x), Domain.greaterThan(w_bank))
                # whether control nonbank industry weight
                if (not (w_nonbank>0 and num_nonbank==0)) and self.industry_control<>0:
                    M.constraint('nonbank', Expr.dot(A_nonbank, x), Domain.greaterThan(w_nonbank))
                # whether control realestate stock weight
                if (not (w_realestate>0 and num_realestate==0)) and self.industry_control==2:
                    M.constraint('realestate', Expr.dot(A_realestate, x), Domain.greaterThan(w_realestate))
                M.constraint('w300', Expr.dot(A_300, x), Domain.lessThan(0.2))
                M.solve()
                sol = x.level()
            for i in xrange(len(sol)):
                if sol[i]<>0:
                    result.append([actual_data[i][0],sol[i]])
                    set_portfolio.append(actual_data[i])
        else:
            # rule for industry control:
            # none stock in target portfolio: add 6 equal weight in portfolio
            # less than 6 in target portfolio: add equal weight until 6 in portfolio
            # more than 6 in target portfolio: scale up weight
            set_portfolio = actual_data[0:self.portfolio_num]
            result = []
            # equal weight in portfolio
            if self.portfolio_weight=='Equal':
                for i in xrange(len(set_portfolio)):
                    result.append([set_portfolio[i][0],1.0/len(set_portfolio)])
            # weighted due to factor
            # note: suppose the factor in set_portfolio all>0 when descending or all<0 when asecending
            #       otherwise need to pre-treat the factor data
            elif self.portfolio_weight=='Factor':
                set_factor = t_factor[0:len(set_portfolio)]
                sum_factor = np.sum(set_factor)
                for i in xrange(len(set_portfolio)):
                    result.append([set_portfolio[i][0],set_factor[i]/sum_factor])
            extra_num = 6
            weightbank = 0
            weightnonbank = 0
            weightrealestate = 0
            weight300 = 0
            weightnot300 = 0
            num_bank = 0
            num_nonbank = 0
            num_realestate = 0
            for j in xrange(len(result)):
                codej = result[j][0]
                pos_code = dict_code[codej]
                if codej[0:3]=='300':# second board
                    weight300 += result[j][1]
                elif indus[pos_code][pos_date-1]=='01030321' and self.industry_control<>0:# bank
                    weightbank += result[j][1]
                    num_bank += 1
                elif indus[pos_code][pos_date-1]=='01030322' and self.industry_control<>0:# nonbank
                    weightnonbank += result[j][1]
                    num_nonbank += 1
                elif indus[pos_code][pos_date-1]=='01030320' and self.industry_control==2:# realestate
                    weightrealestate += result[j][1]
                    num_realestate += 1
                else:
                    weightnot300 += result[j][1]
            if weightbank<w_bank and w_bank<>0 and self.industry_control<>0:
                # none bank stock in target portfolio: add 6 equal weight in portfolio
                if weightbank==0:
                    num = 0           
                    for j in xrange(len(actual_data)):
                        codej = actual_data[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030321':
                            num += 1
                            set_portfolio.append(actual_data[j])
                            result.append([codej,w_bank/extra_num])
                        if num== extra_num:
                            break
                # less than 6 bank stock in target portfolio: add to 6 equal weight stocks in portfolio
                elif num_bank<extra_num:
                    num = 0  
                    code = []
                    for j in xrange(len(result)):
                        codej = result[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030321':
                            num += 1
                            result[j][1] = w_bank/extra_num  
                            code.append(codej)
                    for j in xrange(len(actual_data)):
                        codej = actual_data[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030321' and (codej not in code):
                            num += 1
                            set_portfolio.append(actual_data[j])
                            result.append([codej,w_bank/extra_num])  
                            if num== extra_num:
                                break
                # more than 6 bank stock in target portfolio: scale up weight
                else:
                    for j in xrange(len(result)):
                        codej = result[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030321':
                            result[j][1] = result[j][1]*w_bank/weightbank        
            if weightnonbank<w_nonbank and w_nonbank<>0 and self.industry_control<>0:
                # none nonbank stock in target portfolio: add 6 equal weight in portfolio
                if weightnonbank==0:
                    num = 0           
                    for j in xrange(len(actual_data)):
                        codej = actual_data[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030322':
                            num += 1
                            set_portfolio.append(actual_data[j])
                            result.append([codej,w_nonbank/extra_num])
                        if num== extra_num:
                            break   
                # less than 6 nonbank stock in target portfolio: add to 6 equal weight stocks in portfolio
                elif num_nonbank<extra_num:
                    num = 0  
                    code = []
                    for j in xrange(len(result)):
                        codej = result[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030322':
                            num += 1
                            result[j][1] = w_nonbank/extra_num  
                            code.append(codej)
                    for j in xrange(len(actual_data)):
                        codej = actual_data[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030322' and (codej not in code):
                            num += 1
                            set_portfolio.append(actual_data[j])
                            result.append([codej,w_nonbank/extra_num])  
                            if num== extra_num:
                                break
                # more than 6 nonbank stock in target portfolio: scale up weight
                else:
                    for j in xrange(len(result)):
                        codej = result[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030322':
                            result[j][1] = result[j][1]*w_nonbank/weightnonbank     
            if weightrealestate<w_realestate and w_realestate<>0 and self.industry_control==2:
                # none realestate stock in target portfolio: add 6 equal weight in portfolio
                if weightrealestate==0:
                    num = 0           
                    for j in xrange(len(actual_data)):
                        codej = actual_data[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030320':
                            num += 1
                            set_portfolio.append(actual_data[j])
                            result.append([codej,w_realestate/extra_num])
                        if num== extra_num:
                            break     
                # less than 6 realestate stock in target portfolio: add to 6 equal weight stocks in portfolio
                elif num_realestate<extra_num:
                    num = 0  
                    code = []
                    for j in xrange(len(result)):
                        codej = result[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030320':
                            num += 1
                            result[j][1] = w_realestate/extra_num  
                            code.append(codej)
                    for j in xrange(len(actual_data)):
                        codej = actual_data[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030320' and (codej not in code):
                            num += 1
                            set_portfolio.append(actual_data[j])
                            result.append([codej,w_realestate/extra_num])  
                            if num== extra_num:
                                break
                # more than 6 realestate stock in target portfolio: scale up weight
                else:
                    for j in xrange(len(result)):
                        codej = result[j][0]
                        pos_code = dict_code[codej]
                        if indus[pos_code][pos_date-1]=='01030320':
                            result[j][1] = result[j][1]*w_realestate/weightrealestate
            # Second Board weight<0.2: scale up weight
            if weight300>0.2:
                for j in xrange(len(result)):
                    codej = set_portfolio[j][0]
                    pos_code = dict_code[codej]
                    if codej[0:3]=='300':
                        result[j][1] = result[j][1]*0.2/weight300
                    elif indus[pos_code][pos_date-1]<>'01030321' and indus[pos_code][pos_date-1]<>'01030322' and indus[pos_code][pos_date-1]<>'01030320':
                        if self.industry_control==0:
                            result[j][1] = result[j][1]*0.8/weightnot300  
                        elif self.industry_control==1:
                            result[j][1] = result[j][1]*(0.8-max(w_bank,weightbank)-max(w_nonbank,weightnonbank))/weightnot300  
                        elif self.industry_control==2:
                            result[j][1] = result[j][1]*(0.8-max(w_bank,weightbank)-max(w_nonbank,weightnonbank)-max(w_realestate,weightrealestate))/weightnot300  
            else:
                for j in xrange(len(result)):
                    codej = set_portfolio[j][0]
                    pos_code = dict_code[codej]
                    if indus[pos_code][pos_date-1]<>'01030321' and indus[pos_code][pos_date-1]<>'01030322' and indus[pos_code][pos_date-1]<>'01030320':
                        if self.industry_control==1:
                            result[j][1] = result[j][1]*(1-max(w_bank,weightbank)-max(w_nonbank,weightnonbank))/(weight300+weightnot300)  
                        elif self.industry_control==2:
                            result[j][1] = result[j][1]*(1-max(w_bank,weightbank)-max(w_nonbank,weightnonbank)-max(w_realestate,weightrealestate))/(weight300+weightnot300)  
        return result,set_portfolio,ic
    
    
    
    def cal_tradecost(self,pf1,pf2):
        """
        calculate turnover rate: sum of difference between weight of holding portfolio and buying portfolio 
        """
        cost = 0
        code_change = [pf1[i][1] for i in xrange(len(pf1))]
        code = [pf1[i][0] for i in xrange(len(pf1))]
        for i in xrange(len(pf2)):
            codei = pf2[i][0]
            if codei in code:
                for j in xrange(len(code)):
                    if code[j]==codei:
                        break
                code_change[j]=np.abs(code_change[j]-pf2[i][1])
            else:
                code.append(codei)
                code_change.append(pf2[i][1])
        cost = np.sum(code_change)
        return cost
    
    def cal_300(self,pfw):
        """
        calculate second board weight
        """
        w = 0
        for i in xrange(len(pfw)):
            code = pfw[i][0]
            if code[0:3]=='300':
                w += pfw[i][1]
        return w        
     
    def modify_first(self,raw_target_weight,set_portfolio):
        """
        form actual trading portfolio from target portfolio due to price change from preclose to VWAP at first tradeday
        process buying constraint: suppose we buy other stock by ratio if we cannot buy some stocks
        note: the key is we fix target buying shares 
        """
        target_weight = [[y for y in raw_target_weight[i]] for i in xrange(len(raw_target_weight))]
        sum_factor = 0
        for j in xrange(len(target_weight)):
            if set_portfolio[j][2]<>-1:
                target_weight[j][1] = target_weight[j][1]*(1+set_portfolio[j][2])
            sum_factor += target_weight[j][1]
        for j in xrange(len(target_weight)):
            target_weight[j][1] = target_weight[j][1]/sum_factor
        for j in xrange(len(target_weight)):
            if set_portfolio[j][4]==-1:
                target_weight[j][1] = 0
            elif set_portfolio[j][4]==-0:
                target_weight[j][1] = target_weight[j][1]/2
        sum_factor = 0
        for j in xrange(len(target_weight)):
            sum_factor += target_weight[j][1]
        for j in xrange(len(target_weight)):
            target_weight[j][1] = target_weight[j][1]/sum_factor
        for j in xrange(len(set_portfolio)-1,-1,-1):
            if target_weight[j][1]==0:
                target_weight.pop(j)
                set_portfolio.pop(j)
        return target_weight
        
    def modify_close(self,raw_y_weight,raw_yset_portfolio):
        """
        form last tradeday's portfolio after market close from last tradeday's portfolio at VWAP
        note: the key is we fix stock shares 
        """
        y_weight = [[y for y in raw_y_weight[i]] for i in xrange(len(raw_y_weight))]
        yset_portfolio = [[y for y in raw_yset_portfolio[i]] for i in xrange(len(raw_yset_portfolio))]
        sum_factor = 0
        for j in xrange(len(y_weight)):
            if yset_portfolio[j][3]<>-1:
                y_weight[j][1] = y_weight[j][1]*(1+yset_portfolio[j][3])
            sum_factor += y_weight[j][1]
        for j in xrange(len(y_weight)):
            y_weight[j][1] = y_weight[j][1]/sum_factor
        return y_weight
    
    def modify_y(self,raw_y_weight,raw_yset_portfolio_t):
        """
        form last tradeday's portfolio during today's trading time from that after market close last tradeday
        note: the key is we fix stock shares 
        """
        y_weight = [[y for y in raw_y_weight[i]] for i in xrange(len(raw_y_weight))]
        yset_portfolio_t = [[y for y in raw_yset_portfolio_t[i]] for i in xrange(len(raw_yset_portfolio_t))]
        sum_factor = 0
        for j in xrange(len(y_weight)):
            if yset_portfolio_t[j][2]<>-1:
                y_weight[j][1] = y_weight[j][1]*(1+yset_portfolio_t[j][2])
            sum_factor += y_weight[j][1]
        for j in xrange(len(y_weight)):
            y_weight[j][1] = y_weight[j][1]/sum_factor
        return y_weight
    
    
    def modify_daily(self,raw_target_weight,actual_yweight,yset_portfolio_t,set_portfolio):
        """
        calculate actual portfolio from target portfolio
        note: first treat selling constraint, the key is we fix the weight of stocks that cannot be sold
              and then treat buying constraint, the key is we scale up the weight after selling
        """
        target_weight = [[y for y in raw_target_weight[i]] for i in xrange(len(raw_target_weight))]
        # form target trading portfolio from target portfolio due to price change from preclose to VWAP         
        sum_factor = 0
        for j in xrange(len(target_weight)):
            if set_portfolio[j][2]<>-1:
                target_weight[j][1] = target_weight[j][1]*(1+set_portfolio[j][2])
            sum_factor += target_weight[j][1]
        for j in xrange(len(target_weight)):
            target_weight[j][1] = target_weight[j][1]/sum_factor
        # sellnot used to record the weight of stocks that cannot be sold
        sellnot = 0
        # changeif used to record the stock cannot be scaled up
        changeif = [0 for x in target_weight]
        for j in xrange(len(yset_portfolio_t)):
            codej = yset_portfolio_t[j][0]
            # cannot sell
            if yset_portfolio_t[j][5]==-1:
                # Flag used to record whether in today's portfolio
                Flag = True            
                for k in xrange(len(set_portfolio)):
                    if set_portfolio[k][0]==codej:
                        # in today's portfolio and yesterday's weight>target weight, not change
                        if actual_yweight[j][1]>target_weight[k][1]:
                            target_weight[k][1] = actual_yweight[j][1]
                            sellnot += actual_yweight[j][1]
                            changeif[k] = 1
                        Flag = False
                        break
                # not in today's portfolio, add it
                if Flag:
                    set_portfolio.append(yset_portfolio_t[j])
                    target_weight.append([codej,actual_yweight[j][1]])
                    changeif.append(1)
                    sellnot += actual_yweight[j][1]
            # price limit intraday, suppose we can sell half
            elif yset_portfolio_t[j][5]==0:
                Flag = True            
                for k in xrange(len(set_portfolio)):
                    if set_portfolio[k][0]==codej:
                        # in today's portfolio and yesterday's weight>target weight, sell half
                        if actual_yweight[j][1]>target_weight[k][1]:
                            target_weight[k][1] = (target_weight[k][1]+actual_yweight[j][1])/2
                            sellnot += target_weight[k][1]
                            changeif[k] = 1
                        Flag = False
                        break
                # not in today's portfolio, add half
                if Flag:
                    set_portfolio.append(yset_portfolio_t[j])
                    target_weight.append([codej,actual_yweight[j][1]/2])
                    changeif.append(1)
                    sellnot += actual_yweight[j][1]/2
        for j in xrange(len(set_portfolio)):
            codej = set_portfolio[j][0]
            # cannot buy
            if set_portfolio[j][4]==-1 and changeif[j]==0:
                Flag = True            
                for k in xrange(len(yset_portfolio_t)):
                    if yset_portfolio_t[k][0]==codej:
                        # in today's portfolio and yesterday's weight<target weight, not change
                        if actual_yweight[k][1]<target_weight[j][1]:
                            target_weight[j][1] = actual_yweight[k][1]
                        Flag = False
                        break
                # not in today's portfolio, weight = 0
                if Flag:
                    target_weight[j][1]=0
            # price limit intraday, suppose we can buy half
            if set_portfolio[j][4]==0 and changeif[j]==0:
                Flag = True            
                for k in xrange(len(yset_portfolio_t)):
                    if yset_portfolio_t[k][0]==codej:
                        # in today's portfolio and yesterday's weight<target weight, buy half
                        if actual_yweight[k][1]<target_weight[j][1]:
                            target_weight[j][1] = (actual_yweight[k][1]+target_weight[j][1])/2
                        Flag = False
                        break
                # not in today's portfolio, weight = weight/2 
                if Flag:
                    target_weight[j][1]=target_weight[j][1]/2
        # first sell and then buy
        sum_factor = 0
        for j in xrange(len(target_weight)):
            if changeif[j]==0:
                sum_factor += target_weight[j][1]
        for j in xrange(len(target_weight)):
            if changeif[j]==0:
                target_weight[j][1] = target_weight[j][1]*(1-sellnot)/sum_factor
        # delete stock that weight = 0
        for j in xrange(len(set_portfolio)-1,-1,-1):
            if target_weight[j][1]==0:
                target_weight.pop(j)
                set_portfolio.pop(j) 
        # calculate turnover rate
        tc = self.cal_tradecost(target_weight,actual_yweight)    
        return target_weight,tc       
        
        
            
    def backtest(self,startday,endday,base_indus_day):
        """Backtesting core
        Generate backtesting data from day to day
        Note: in the backtesting core, we suppose we get the factor and target portfolio(stock and weight) after a tradeday close
              and then modify the weight dut to nextday's VWAP(target buying quantity fixed) as actual trading portfolio 
              and trade at price VWAP, there may be a little hypothesis error and it's quite close to actual trading
              so the earning one day is divided into preclose to VWAP of last tradeday's portfolio and VWAP to close of today's portfolio
        """
        if (self.index<>'000906') and (self.index<>self.hedge_target):
            print 'Error: index must be consistent with hedge target'
        else:
            Date_Match,Date,dict_date,dict_code,stock_info,codelist,dict_code,is_merger,pct_chg1,pct_chg2,buy_if,sell_if,factor,indus,category_indus,index_info = self.get_backtest_data(startday,endday,base_indus_day)
            if not Date_Match:
                print 'Error: Market and Factor Datatime must be consistent with startday and endday'
            else:
                # flag used to judge the first day to take a position
                flag = True
                # car used to record strategy net value
                car = 1
                # ctr used to record strategy turnover rate
                ctr = 0
                # cic used to record ic
                cic = 0
                for i in xrange(len(Date)):
                    tdate = Date[i]
                    pos_date = i
                    stock_pool = []
                    ar = 0
                    for j in xrange(len(stock_info)):
                        pos_code = dict_code[stock_info[j][0]]
                        # index consitution stock and not get merged
                        if stock_info[j][1]<=tdate and stock_info[j][2]>tdate and (is_merger[pos_code][0]==0 or (is_merger[pos_code][0]==1 and is_merger[pos_code][1]>=i)):
                            stock_pool.append([stock_info[j][0]])
                    # num_qualify records the num of stocks which factor is effective 
                    num_qualify = 0
                    # select stock that factor<>-1 where -1 means nan in factor data
                    for j in xrange(len(stock_pool)):
                        codej = stock_pool[j][0]
                        pos_code = dict_code[codej]
                        t_factor = factor[pos_code][pos_date]
                        t_pct1 = pct_chg1[pos_code][pos_date]
                        t_pct2 = pct_chg2[pos_code][pos_date]
                        t_buyif = buy_if[pos_code][pos_date]
                        t_sellif = sell_if[pos_code][pos_date]
                        if t_factor!=-1:
                            stock_pool[j].append(t_factor)
                            stock_pool[j].append(t_pct1)
                            stock_pool[j].append(t_pct2)
                            stock_pool[j].append(t_buyif)
                            stock_pool[j].append(t_sellif)
                            num_qualify += 1
                    # suppose we take position when qualified stock num more than 4*portfolio_num 
                    if num_qualify<self.portfolio_num*4:
                        continue
                    else:
                        actual_data = []
                        # select qualified stock that there is not continuous trade suspension in 40 tradedays 
                        for j in xrange(len(stock_pool)):
                            day_sus = 0
                            codej = stock_pool[j][0]
                            pos_code = dict_code[codej]
                            for k in xrange(i-40,i):
                                if buy_if[pos_code][k]<>1 and buy_if[pos_code][k+1]<>1:
                                    day_sus += 1
                            if len(stock_pool[j])==6 and day_sus<20:
                                actual_data.append(stock_pool[j])
                        w_bank = max(0,index_info[i-1][2]/100-self.weight_low_bank)
                        w_nonbank = max(0,index_info[i-1][3]/100-self.weight_low_nonbank)
                        w_realestate = max(0,index_info[i-1][4]/100-self.weight_low_realestate)
                        # industry neutralize
                        if self.industry_neutral==1:
                            # Last tradeday's industry category
                            for j in xrange(len(category_indus[i-1])):
                                position = []
                                x = []
                                for k in xrange(len(actual_data)):
                                    codek = actual_data[k][0]
                                    pos_code = dict_code[codek]
                                    if indus[pos_code][pos_date-1]==category_indus[i-1][j]:
                                        position.append(k)
                                        x.append(actual_data[k][1])
                                mid = np.median(x)
                                for k in xrange(len(position)):
                                    actual_data[position[k]][1] = actual_data[position[k]][1]-mid    
                        if self.factor_order==1:
                            actual_data.sort(key=lambda x:x[1],reverse=True)
                        else:
                            actual_data.sort(key=lambda x:x[1])
                        # calculate target portfolio
                        target_weight,set_portfolio,ic = self.cal_tweight(actual_data,w_bank,w_nonbank,w_realestate,indus,i,pct_chg1,pct_chg2,Date,dict_code)
                        if flag:
                            # modify_first: from target_portfolio to actual trading portfolio on first tradeday
                            actual_weight = self.modify_first(target_weight,set_portfolio)  
                            weight300 = self.cal_300(actual_weight)
                            for j in xrange(len(actual_weight)):
                                if set_portfolio[j][3]<>-1:
                                    ar += (1+set_portfolio[j][3])*actual_weight[j][1]
                            ar = ar -1
                            tc = 1
                            start_date = i
                            flag = False
                            y_weight = [[y for y in actual_weight[j]] for j in xrange(len(actual_weight))]
                            yset_portfolio = [[y for y in set_portfolio[j]] for j in xrange(len(set_portfolio))]
                        else:
                            # yset_portfolio_t used as yesterday's portfolio's market data today
                            yset_portfolio_t = []
                            for j in xrange(len(y_weight)):
                                codej = y_weight[j][0]
                                pos_code = dict_code[codej]
                                t_factor = factor[pos_code][pos_date]
                                t_pct1 = pct_chg1[pos_code][pos_date]
                                t_pct2 = pct_chg2[pos_code][pos_date]
                                t_buyif = buy_if[pos_code][pos_date]
                                t_sellif = sell_if[pos_code][pos_date]
                                yset_portfolio_t.append([codej,t_factor,t_pct1,t_pct2,t_buyif,t_sellif])
                            # modify_close:from actual portfolio intraday to portfolio at close price 
                            close_yweight = self.modify_close(y_weight,yset_portfolio)
                            pct1 = 0
                            for j in xrange(len(close_yweight)):
                                if yset_portfolio_t[j][2]<>-1:
                                    pct1 += (yset_portfolio_t[j][2]+1)*close_yweight[j][1]
                                else:
                                    pct1 += close_yweight[j][1]       
                            # modify_y: from yesterday's portfolio at close price to today's VWAP, used to calculate turnover rate
                            actual_yweight = self.modify_y(close_yweight,yset_portfolio_t)
                            # modify_daily: from target portfolio to actual portfolio, considering buy and sell constrain
                            actual_weight,tc = self.modify_daily(target_weight,actual_yweight,yset_portfolio_t,set_portfolio)
                            weight300 = self.cal_300(actual_weight)
                            pct2 = 0
                            for j in xrange(len(actual_weight)):
                                if set_portfolio[j][3]<>-1:
                                    pct2 += (1+set_portfolio[j][3])*actual_weight[j][1]
                                else:
                                    pct2 += actual_weight[j][1]
                            ar = pct1 * pct2- 1 
                        for j in xrange(len(index_info)):
                            if index_info[j][0]==Date[i]:
                                index_tar = index_info[j][1]
                                break
                        car = car + 0.5 * (ar - index_tar - tc*self.trade_cost)
                        ctr = ctr + tc*100
                        cic = cic + ic
                        y_weight = [[y for y in actual_weight[j]] for j in xrange(len(actual_weight))]
                        yset_portfolio = [[y for y in set_portfolio[j]] for j in xrange(len(set_portfolio))]
                        self.daily_excess_return.append(0.5*(ar-index_tar-tc*0.0012))
                        self.turnover_rate.append(ctr)
                        self.actual_portfolio_num.append(len(actual_weight))
                        self.weight_300.append(weight300)
                        self.rank_ic.append(cic)
                        self.net_value.append(car)
                        print Date[i]
                self.trade_date = Date[start_date:len(Date)]
          
          
          
    def analyze(self,startday):
        """
        analyaze the strategy backtesting performance
        note: include turnover rate, stock numbers, rank ic and weight of GEM for chart
              include annualized return, max drawdown, sharpe ratio, winning rate and P2L rate for table
        """
        path1=u'.\\backtest.pdf'
        path2=u'.\\backtest.xlsx'
        pdate = self.trade_date
        pp = PdfPages(path1) 
        plt.figure()
        plt.clf()            
        ax=plt.gca()
        ax.set_xticks(np.linspace(0,len(pdate),10))
        ax.set_xticklabels((pdate[0], pdate[len(pdate)/9], pdate[2*len(pdate)/9], pdate[3*len(pdate)/9],\
        pdate[4*len(pdate)/9], pdate[5*len(pdate)/9],pdate[6*len(pdate)/9],pdate[7*len(pdate)/9],\
        pdate[8*len(pdate)/9],pdate[len(pdate)-1]))
        for label in ax.xaxis.get_ticklabels():
            label.set_rotation(45)
            label.set_fontsize(6)
        plt.plot(self.net_value,color='b')
        plt.xlabel
        pp.savefig()
        plt.figure()
        plt.clf()   
        plt.subplot(221)   
        # cumulative turnover rate
        tr = self.turnover_rate
        plt.plot(tr)
        plt.xticks(fontsize=8)
        plt.title('turnover rate average='+'%.2f'%(tr[len(tr)-1]/len(tr)),fontsize=8)
        plt.subplot(222)   
        # numbers of stocks in the portfolio
        nport = self.actual_portfolio_num
        plt.plot(nport)
        plt.xticks(fontsize=8)
        plt.title('portfolio num average='+'%.2f'%(np.mean(nport)),fontsize=8)
        plt.subplot(223)  
        # cumulative rank ic
        ic = self.rank_ic
        plt.plot(ic)
        plt.xticks(fontsize=8)
        titlenum = ic[len(ic)-1]/len(ic)*100
        plt.title('rank ic average='+'%.2f'%(titlenum),fontsize=8)
        plt.subplot(224)  
        # weight of second board in the portfolio
        w300 = self.weight_300
        plt.plot(w300)
        plt.xticks(fontsize=8)
        k = 0
        while w300[k]==0:
            k = k + 1
        w300 = w300[k:len(w300)]
        titlenum = np.mean(w300)*100
        plt.title('weight of 300 average='+'%.2f'%(titlenum),fontsize =8)
        pp.savefig()
        pp.close()
        year_time = []
        year_start = startday[0:4]
        # generate year=[year, first trade date, last trade date]
        for i in xrange(len(pdate)):
            year = pdate[i][0:4]
            if year_time == []:
                year_time.append([year,i,0])
                year_start = year
            elif year !=year_start:
                year_time[len(year_time)-1][2]=i-1
                year_time.append([year,i,0])
                year_start = year
            elif i==len(pdate)-1:
                year_time[len(year_time)-1][2]=i
        strategy_valuate = [] 
        for i in xrange(len(year_time)):
            # calculate some statistics by year 
            der = self.daily_excess_return[year_time[i][1]:year_time[i][2]+1]
            day = pdate[year_time[i][1]:year_time[i][2]+1]
            row = self.valuate(der,day)
            strategy_valuate.append(row)
        row = self.valuate(self.daily_excess_return,pdate)
        strategy_valuate.append(row)
        index = ['Year','Return','MDD','MDD_SPAN','Sharpe','Daily_Min','Min_Day','Win_Ratio','P2L_Ratio']
        wb = Workbook()
        ew = ExcelWriter(workbook=wb)
        ws = wb.worksheets[0]
        for i in xrange(len(index)):
            ws.cell(row=1,column=i+1).value = index[i]
        for i in xrange(len(year_time)):
            ws.cell(row=i+2,column=1).value = year_time[i][0]
        ws.cell(row=len(year_time)+2,column=1).value = u'年化'
        for i in xrange(len(strategy_valuate)):
            for j in xrange(len(strategy_valuate[i])):
                ws.cell(row=i+2,column=j+2).value = strategy_valuate[i][j]
        ew.save(filename=path2)  
        
    
    